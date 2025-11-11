from rest_framework.decorators import api_view
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, NumberFormatDescriptor
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import *
from .serializer import *
import decimal
from django.db.models import Sum
from datetime import date, datetime,time, timedelta
import json
import logging

# Configurar logging para depuración
logger = logging.getLogger(__name__)

@api_view(['POST'])
def excel_regmani(request):
    # Parsear el cuerpo de la solicitud
    try:
        vue_data = json.loads(request.body) if request.body else {}
        logger.info("Datos recibidos: %s", vue_data)  # Para depuración
    except json.JSONDecodeError:
        logger.error("Error al parsear JSON: %s", request.body)
        return Response({"error": "Invalid JSON data"}, status=400)

    # Validar que todos los campos de nivel raíz estén presentes
    required_fields = [
        "operador", "aeronave_matricula", "puerto_despacho",
        "puerto_destino", "fecha", "no_vuelo", "tabla"
    ]
    missing_fields = [field for field in required_fields if field not in vue_data or vue_data[field] is None]
    if missing_fields:
        logger.error("Campos obligatorios faltantes: %s", missing_fields)
        return Response({"error": f"Missing required fields: {', '.join(missing_fields)}"}, status=400)

    # Extraer los campos
    operador = vue_data["operador"]
    aeronave_matricula = vue_data["aeronave_matricula"]
    puerto_despacho = vue_data["puerto_despacho"]
    puerto_destino = vue_data["puerto_destino"]
    fecha = vue_data["fecha"]
    no_vuelo = vue_data["no_vuelo"]
    tabla = vue_data["tabla"]

    # Validar que tabla sea un arreglo
    if not isinstance(tabla, list):
        logger.error("El campo 'tabla' debe ser un arreglo")
        return Response({"error": "'tabla' must be an array"}, status=400)

    # Validar que cada elemento de tabla tenga los campos requeridos (excepto destino y peso)
    required_tabla_fields = ["guia", "items", "naturaleza", "destinatario"]
    for index, item in enumerate(tabla):
        missing_tabla_fields = [field for field in required_tabla_fields if field not in item or item[field] is None]
        if missing_tabla_fields:
            logger.error("Campos faltantes en el elemento %d de 'tabla': %s", index, missing_tabla_fields)
            return Response({"error": f"Missing fields in tabla item {index}: {', '.join(missing_tabla_fields)}"}, status=400)

    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Manifiesto de Carga"

    id_compania = '300'
    id_agencia = '001'
    # Set up fonts and alignment
    center_alignment = Alignment(horizontal="center")

    # Sección del encabezado
    ws.merge_cells('A1:G1')
    ws['A1'] = "CARGO MANIFEST"
    ws['A1'].font = Font(name="Arial", size=16, bold=True)
    ws['A1'].alignment = center_alignment

    ws.merge_cells('A2:G2')
    ws['A2'] = "MANIFIESTO DE CARGA"
    ws['A2'].font = Font(name="Arial", size=18, bold=False)
    ws['A2'].alignment = center_alignment

    ws['A4'] = "OWNER OR OPERATOR"
    ws['A4'].font = Font(name="Arial", size=8, bold=False)

    ws.merge_cells('C4:F4')
    ws['C4'] = operador
    ws['C4'].font = Font(name="Arial", size=14, bold=True)
    ws['C4'].alignment = center_alignment

    ws['A5'] = "Propietario o Administrador"
    ws['A5'].font = Font(name="Arial", size=8, bold=False)

    ws['A6'] = "AIRCRAFT"
    ws['A6'].font = Font(name="Arial", size=6, bold=True)

    ws['E6'] = "MARKS OF NATIONALITY AND REGISTRATION"
    ws['E6'].font = Font(name="Arial", size=6, bold=False)

    ws['G6'] = "FLIGHT NO."
    ws['G6'].font = Font(name="Arial", size=6, bold=False)

    ws['A7'] = "Aeronave"
    ws['A7'].font = Font(name="Arial", size=6, bold=False)

    ws['E8'] = "MARCAS DE NACIONALIDAD Y MATRICULA"
    ws['E8'].font = Font(name="Arial", size=6, bold=False)

    ws['G8'] = "VUELO NO."
    ws['G8'].font = Font(name="Arial", size=6, bold=False)

    ws['A9'] = operador.split()[0]
    ws['A9'].font = Font(name="Arial", size=20, bold=True)

    ws.merge_cells('E9:F9')
    ws['E9'] = aeronave_matricula
    ws['E9'].font = Font(name="Arial", size=20, bold=True)

    ws['G9'] = no_vuelo
    ws['G9'].font = Font(name="Arial", size=20, bold=True)

    ws['A10'] = "POINT OF LOADING"
    ws['A10'].font = Font(name="Arial", size=6, bold=False)

    ws['E10'] = "POINT OF UNLOADING"
    ws['E10'].font = Font(name="Arial", size=6, bold=False)

    ws['G10'] = "DATE"
    ws['G10'].font = Font(name="Arial", size=6, bold=False)

    ws['A11'] = "PUNTO DE EMBARQUE"
    ws['A11'].font = Font(name="Arial", size=6, bold=False)

    ws['E11'] = "PUNTO DE DESCARGA"
    ws['E11'].font = Font(name="Arial", size=6, bold=False)

    ws['G11'] = "FECHA"
    ws['G11'].font = Font(name="Arial", size=8, bold=False)

    ws.merge_cells('A12:B12')
    ws['A12'] = puerto_despacho
    ws['A12'].font = Font(name="Arial", size=20, bold=True)

    ws.merge_cells('E12:F12')
    ws['E12'] = puerto_destino
    ws['E12'].font = Font(name="Arial", size=20, bold=True)

    ws['G12'] = fecha
    ws['G12'].font = Font(name="Arial", size=20, bold=True)

    # Encabezado de la tabla
    ws['A13'] = "AIR WAYBILL"
    ws['A13'].font = Font(name="Arial", size=6, bold=False)
    ws['A13'].alignment = center_alignment

    ws['B13'] = "NUMBER"
    ws['B13'].font = Font(name="Arial", size=6, bold=False)
    ws['B13'].alignment = center_alignment

    ws['C13'] = "WEIGHT"
    ws['C13'].font = Font(name="Arial", size=6, bold=False)
    ws['C13'].alignment = center_alignment

    ws['D13'] = "DESTINO"
    ws['D13'].font = Font(name="Arial", size=6, bold=False)
    ws['D13'].alignment = center_alignment

    ws['E13'] = "NATURE OF GOODS"
    ws['E13'].font = Font(name="Arial", size=6, bold=False)
    ws['E13'].alignment = center_alignment

    ws['F13'] = "CONSIGNEE AND ADDRESS"
    ws['F13'].font = Font(name="Arial", size=6, bold=False)
    ws['F13'].alignment = center_alignment

    ws['A14'] = "NUMBER"
    ws['A14'].font = Font(name="Arial", size=6, bold=False)
    ws['A14'].alignment = center_alignment

    ws['B14'] = "OF PIECES"
    ws['B14'].font = Font(name="Arial", size=6, bold=False)
    ws['B14'].alignment = center_alignment

    ws['C14'] = "KGS"
    ws['C14'].font = Font(name="Arial", size=6, bold=False)
    ws['C14'].alignment = center_alignment

    # Tabla
    start_data_row = 17
    for i, item in enumerate(tabla, start=start_data_row):
        try:
            ws.cell(row=i, column=1, value=item["guia"])
            ws.cell(row=i, column=2, value=item["items"] if item["items"] is not None else 0)
            # Asignar valor por defecto a peso si es inválido
            peso = item.get("peso", "0")
            if not peso.strip() or not peso.replace(".", "").isdigit():
                peso = "0"
            ws.cell(row=i, column=3, value=float(peso))
            ws.cell(row=i, column=4, value=item.get("destino", "PTY"))
            ws.cell(row=i, column=5, value=item["naturaleza"])
            ws.cell(row=i, column=6, value=item["destinatario"])
        except (KeyError, ValueError) as e:
            logger.error("Error al procesar elemento %d de tabla: %s", i - start_data_row, str(e))
            return Response({"error": f"Invalid data in tabla item {i - start_data_row}: {str(e)}"}, status=400)

    # Adjust column widths
    column_widths = [20, 6.14, 8.71, 4.57, 35.86, 41.57, 21.14]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    id_passd = Segpaag.objects.get(compania=id_compania, agencia=id_agencia, aplicacion='CAR',
                            parametro='excel_passd').valor
    
    ws.protection.password = id_passd.strip()
    ws.protection.sheet = True
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="cargo_manifest.xlsx"'
    wb.save(response)
    return response

@api_view(['POST'])
def excel_caratvue(request):
    # Parsear el cuerpo de la solicitud
    try:
        vue_data = json.loads(request.body) if request.body else {}
        logger.info("Datos recibidos: %s", vue_data)  # Para depuración
    except json.JSONDecodeError:
        logger.error("Error al parsear JSON: %s", request.body)
        return Response({"error": "Invalid JSON data"}, status=400)

    # Validar que todos los campos de nivel raíz estén presentes
    required_fields = [
        "fecha", "compania", "matricula",
        "aeronave", "fecha_llegada", "hora_llegada", "status", "tabla" ]
    missing_fields = [field for field in required_fields if field not in vue_data or vue_data[field] is None]
    if missing_fields:
        logger.error("Campos obligatorios faltantes: %s", missing_fields)
        return Response({"error": f"Missing required fields: {', '.join(missing_fields)}"}, status=400)

    # Extraer los campos
    rfecha = vue_data["fecha"]
    rcompania = vue_data["compania"]
    rmatricula = vue_data["matricula"]
    id_aeronave = vue_data["aeronave"]
    fecha_llegada = vue_data["fecha_llegada"]
    hora_llegada = vue_data["hora_llegada"]
    estado = vue_data["status"]
    tabla = vue_data["tabla"]

    # Validar que tabla sea un arreglo
    if not isinstance(tabla, list):
        logger.error("El campo 'tabla' debe ser un arreglo")
        return Response({"error": "'tabla' must be an array"}, status=400)

    # Validar que cada elemento de tabla tenga los campos requeridos (excepto destino y peso)
    required_tabla_fields = ["cargo", "tiempo_total", "monto", "status"]
    for index, item in enumerate(tabla):
        missing_tabla_fields = [field for field in required_tabla_fields if field not in item or item[field] is None]
        if missing_tabla_fields:
            logger.error("Campos faltantes en el elemento %d de 'tabla': %s", index, missing_tabla_fields)
            return Response({"error": f"Missing fields in tabla item {index}: {', '.join(missing_tabla_fields)}"}, status=400)

    now = datetime.now()
    id_fecha = now.date()
    w_hora = now.time()
    id_hora = w_hora.strftime('%H:%M:%S')
    id_compania = '300'
    id_agencia = '001'

    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Servicios a Aeronaves"

    # Set up fonts and alignment
    center_alignment = Alignment(horizontal="center")
    right_alignment = Alignment(horizontal="right")
    left_alignment = Alignment(horizontal="left")

    tamano12 = Font(size=12)
    tamano13 = Font(size=13)
    negrita12 = Font(size=12, bold=True)
    negrita13 = Font(size=13, bold=True)

    # Sección del encabezado
    ws.merge_cells('C1:E1')
    ws['C1'] = "COPADASA"
    ws['C1'].font = Font(name="Arial", size=16, bold=False)
    ws['C1'].alignment = center_alignment

    ws.merge_cells('C2:E2')
    ws['C2'] = "AUXILIAR SERVICIOS AERONAVES"
    ws['C2'].font = Font(name="Arial", size=16, bold=True)
    ws['C2'].alignment = center_alignment

    ws['A1'] = "Fecha:"
    ws['A1'].font = Font(name="Arial", size=11, bold=False)
    ws['A1'].alignment = right_alignment
     
    ws['B1'] = id_fecha
    ws['B1'].font = Font(name="Arial", size=11, bold=False)
    ws['B1'].alignment = left_alignment

    ws['A4'] = "Fecha:"
    ws['A4'].font = Font(name="Arial", size=13, bold=True)
    ws['A4'].alignment = right_alignment
    ws['B4'] = rfecha
    ws['B4'].font = Font(name="Arial", size=13, bold=True)

    ws['E4'] = "Estado:"
    ws['E4'].font = Font(name="Arial", size=11, bold=False)
    ws['E4'].alignment = right_alignment

    if estado == "R":
        estado = "Registrado"
    if estado == "A":
        estado = "Anulado"
    if estado == "C":
        estado = "Cerrado"

    ws['F4'] = estado
    ws['F4'].font = Font(name="Arial", size=11, bold=False)

    ws['A2'] = "Hora:"
    ws['A2'].font = Font(name="Arial", size=11, bold=False)
    ws['A2'].alignment = right_alignment
    ws['B2'] = id_hora
    ws['B2'].font = Font(name="Arial", size=11, bold=False)
  
    ws['A5'] = "Compañía:"
    ws['A5'].font = Font(name="Arial", size=13, bold=False)
    ws['A5'].alignment = right_alignment
    ws['B5'] = rcompania
    ws['B5'].font = Font(name="Arial", size=13, bold=False)
    ws['B5'].alignment = center_alignment

    ncompania = Carcoaer.objects.get(compania=rcompania).nombre
    ws['C5'] = ncompania
    ws['C5'].font = Font(name="Arial", size=13, bold=False)

    ws['E5'] = "Fecha Llegada:"
    ws['E5'].font = Font(name="Arial", size=13, bold=False)
    ws['E5'].alignment = right_alignment
    ws['F5'] = fecha_llegada
    ws['F5'].font = Font(name="Arial", size=13, bold=False)

    ws['A6'] = "Matrícula:"
    ws['A6'].font = Font(name="Arial", size=13, bold=False)
    ws['A6'].alignment = right_alignment
    ws['B6'] = rmatricula
    ws['B6'].font = Font(name="Arial", size=13, bold=False)

    ws['E6'] = "Hora Llegada:"
    ws['E6'].font = Font(name="Arial", size=13, bold=False)
    ws['E6'].alignment = right_alignment
    ws['F6'] = hora_llegada
    ws['F6'].font = Font(name="Arial", size=13)

    
    naeronave = Cartiaero.objects.get(aeronave=id_aeronave).descripcion

    ws['A7'] = "Aeronave"
    ws['A7'].font = Font(name="Arial", size=13, bold=False)
    ws['B7'] = id_aeronave
    ws['B7'].font = Font(name="Arial", size=13, bold=False)
    ws['B7'].alignment = center_alignment
    ws['C7'] = naeronave
    ws['C7'].font = Font(name="Arial", size=13, bold=False)

    resultado = Caratvued.objects.filter(fecha=rfecha, compania=rcompania, matricula=rmatricula).aggregate(total=Sum('monto'))
    monto = decimal.Decimal(resultado['total'])

    ws['D7'] = "Total:"
    ws['D7'].font = Font(name="Arial", size=13, bold=False)
    ws['D7'].alignment = right_alignment
    ws['E7'] = float(monto)
    ws['E7'].font = Font(name="Arial", size=13, bold=True)
    ws['E7'].number_format = '$###,##0.00'

    # Encabezado de la tabla
    ws['A9'] = "CARGO"
    ws['A9'].font = Font(name="Arial", size=10, bold=True)
    ws['A9'].alignment = center_alignment

    ws['B9'] = "DESCRIPCIÓN"
    ws['B9'].font = Font(name="Arial", size=10, bold=True)
    ws['B9'].alignment = left_alignment

    ws['C9'] = "TIEMPO"
    ws['C9'].font = Font(name="Arial", size=10, bold=True)
    ws['C13'].alignment = center_alignment

    ws['D9'] = "MONTO"
    ws['D9'].font = Font(name="Arial", size=10, bold=True)
    ws['D9'].alignment = right_alignment

    ws['E9'] = "ESTADO"
    ws['E9'].font = Font(name="Arial", size=10, bold=True)
    ws['E9'].alignment = center_alignment

    # Tabla
    totMonto = 0
    start_data_row = 10
    for i, item in enumerate(tabla, start=start_data_row):
        try:
            ws.cell(row=i, column=1, value=item["cargo"]).alignment = center_alignment
            #ws.cell(row=i, column=2, value=item["items"] if item["items"] is not None else 0)
            # Asignar valor por defecto a peso si es inválido
            #peso = item.get("peso", "0")
            #if not peso.strip() or not peso.replace(".", "").isdigit():
            #    peso = "0"
            if item["status"] == "R":
                nstatus = "Registrado"
            if item["status"] == "C":
                nstatus = "Cerrado"
            if item["status"] == "A":
                nstatus = "Anulado"

            wmonto = item["monto"]
            wmonto = float(wmonto)
            totMonto = totMonto + wmonto
            wcargo = item["cargo"]

            ncargo = Caratenvue.objects.get(cargo=wcargo).nombre
            ws.cell(row=i, column=2, value=ncargo).font = tamano12
            ws.cell(row=i, column=3, value=item["tiempo_total"]).alignment = center_alignment
            ws.cell(row=i, column=3, value=item["tiempo_total"]).font = tamano12
            ws.cell(row=i, column=4, value=wmonto).alignment = right_alignment
            ws.cell(row=i, column=4, value=wmonto).number_format = "$###,##0.00"
            ws.cell(row=i, column=4, value=wmonto).font = tamano12
            ws.cell(row=i, column=5, value=nstatus).alignment = center_alignment
            ws.cell(row=i, column=5, value=nstatus).font = tamano12
             
        except (KeyError, ValueError) as e:
            logger.error("Error al procesar elemento %d de tabla: %s", i - start_data_row, str(e))
            return Response({"error": f"Invalid data in tabla item {i - start_data_row}: {str(e)}"}, status=400)

    i+= 1
    ws.cell(row=i, column=3, value="Total:").alignment = right_alignment
    ws.cell(row=i, column=3, value="Total:").font = negrita12
    ws.cell(row=i, column=4, value=totMonto).alignment = right_alignment 
    ws.cell(row=i, column=4, value=totMonto).number_format = "$###,##0.00"
    ws.cell(row=i, column=4, value=totMonto).font = negrita12

    i+= 2
    ws.cell(row=i, column=1, value="Registrado Por:")
    ws.cell(row=i, column=2, value="contesis")
    i+= 1
    ws.cell(row=i, column=1, value="View:")
    ws.cell(row=i, column=2, value="excel_caratvue")

    # Adjust column widths
    column_widths = [12.85, 14.91, 19.18, 12.36, 17.30, 12.55]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    id_passd = Segpaag.objects.get(compania=id_compania, agencia=id_agencia, aplicacion='CAR',
                            parametro='excel_passd').valor
    
    ws.protection.password = id_passd.strip()
    ws.protection.sheet = True
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="servicios_aereos.xlsx"'
    wb.save(response)
    return response

@api_view(['POST'])
def excel_logctmo(request):
    # Parsear el cuerpo de la solicitud
    try:
        vue_data = json.loads(request.body) if request.body else {}
        logger.info("Datos recibidos: %s", vue_data)  # Para depuración
    except json.JSONDecodeError:
        logger.error("Error al parsear JSON: %s", request.body)
        return Response({"error": "Invalid JSON data"}, status=400)

    # Validar que todos los campos de nivel raíz estén presentes
    required_fields = [
        "fecha_desde", "fecha_hasta" ]
    missing_fields = [field for field in required_fields if field not in vue_data or vue_data[field] is None]
    if missing_fields:
        logger.error("Campos obligatorios faltantes: %s", missing_fields)
        return Response({"error": f"Missing required fields: {', '.join(missing_fields)}"}, status=400)

    # Extraer los campos
    fecha_desde = vue_data["fecha_desde"]
    fecha_hasta = vue_data["fecha_hasta"]
    codigo = vue_data["codigo"]
    cliente = vue_data["cliente"]
    documento = vue_data["documento"]
    estado = vue_data["status"]

    id_compania = '300'
    id_agencia = '001'
    id_almacen = '02'
     
    now = datetime.now()
    id_fecha = now.date()
    w_hora = now.time()
    id_hora = w_hora.strftime('%H:%M:%S')

    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Servicios a Aeronaves"

    # Set up fonts and alignment
    center_alignment = Alignment(horizontal="center")
    right_alignment = Alignment(horizontal="right")

    # Sección del encabezado
    ws.merge_cells('A1:K1')
    ws['A1'] = "COPADASA"
    ws['A1'].font = Font(name="Arial", size=16, bold=False)
    ws['A1'].alignment = center_alignment

    ws.merge_cells('A2:K2')
    ws['A2'] = "MOVIMIENTOS DE CUARTO FRÍO"
    ws['A2'].font = Font(name="Arial", size=18, bold=True)
    ws['A2'].alignment = center_alignment

    ws['D3'] = "DESDE:"
    ws['D3'].font = Font(name="Arial", size=14, bold=True)
    ws['D3'].alignment = right_alignment
    ws['E3'] = fecha_desde
    ws['E3'].font = Font(name="Arial", size=14, bold=True)

    ws['G3'] = "HASTA:"
    ws['G3'].font = Font(name="Arial", size=14, bold=True)
    ws['G3'].alignment = right_alignment
    ws['H3'] = fecha_hasta
    ws['H3'].font = Font(name="Arial", size=14, bold=True)

    ws['A4'] = "Fecha:"
    ws['A4'].font = Font(name="Arial", size=13, bold=False)
    ws['A4'].alignment = right_alignment

    ws['B4'] = id_fecha
    ws['B4'].font = Font(name="Arial", size=13, bold=False)

    if estado:
        ws['J4'] = "Estado:"
        ws['J4'].font = Font(name="Arial", size=13, bold=False)
        ws['J4'].alignment = right_alignment

        if estado == "I":
            estado = "Ingresado"
        if estado == "C":
            estado = "Calculado"
        if estado == "E":
            estado = "Despachado"

        ws['I4'] = estado
        ws['I4'].font = Font(name="Arial", size=13, bold=False)
    
    ws['A5'] = "Hora:"
    ws['A5'].font = Font(name="Arial", size=13, bold=False)
    ws['A5'].alignment = right_alignment
    ws['B5'] = id_hora
    ws['B5'].font = Font(name="Arial", size=13, bold=False)

    sw = '0'
    logctmo = Logctmo.objects.filter(compania=id_compania, agencia=id_agencia, 
                fecha__gte=fecha_desde, fecha__lte=fecha_hasta, almacen=id_almacen).values().order_by('fecha', 'codigo', 'cliente', 'documento')
  
    tamano12 = Font(size=12)
    tamano13 = Font(size=13)
    negrita12 = Font(size=12, bold=True)
    negrita13 = Font(size=13, bold=True)
    i = 7
    sw = '0'
    for ctmo in logctmo:
        wcompania = ctmo['compania']
        wagencia = ctmo['agencia']
        wfecha = ctmo['fecha']
        walmacen = ctmo['almacen']
        wcodigo = ctmo['codigo']
        wdocumento = ctmo['documento']
        wcliente = ctmo['cliente']
        wfechaEntrega = ctmo['fecha_entrega']
        whoraEntrega = ctmo['hora_entrega']
        wguia = ctmo['guia_despacho']
        wfechaLlegada = ctmo['fecha_llegada']
        whoraLlegada = ctmo['hora_llegada']
        wmonto = ctmo['valor']
        wstatus = ctmo['status']
        wcomentario = ctmo['comentario']

        if sw == '0':
            nalmacen = Logalma.objects.get(almacen=walmacen).descripcion
            ncodigo = Logtral.objects.get(codigo=wcodigo).descripcion
            ws.cell(row=i, column=1, value="Fecha:").alignment = right_alignment
            ws.cell(row=i, column=1, value="Fecha:").font = negrita13
            ws.cell(row=i, column=2, value=wfecha).alignment = center_alignment
            ws.cell(row=i, column=2, value=wfecha).font = negrita13
            i+= 1
            ws.cell(row=i, column=1, value="Almacén:").alignment = right_alignment
            ws.cell(row=i, column=1, value="Almacen:").font = tamano13
            ws.cell(row=i, column=2, value=nalmacen)
            ws.cell(row=i, column=2, value=nalmacen).font = tamano13
            i+= 1
            ws.cell(row=i, column=1, value="Transacción:").alignment = right_alignment
            ws.cell(row=i, column=1, value="Transacción:").font = tamano13
            ws.cell(row=i, column=2, value=ncodigo).font = tamano13
            i+=2

            ws.cell(row=i, column=6, value="Fecha").font = negrita12
            ws.cell(row=i, column=7, value="Hora").font = negrita12
            ws.cell(row=i, column=8, value="Fecha").font = negrita12
            ws.cell(row=i, column=9, value="Hora").font = negrita12
            i+= 1
            ws.cell(row=i, column=1, value="Documento").font = negrita12
            ws.cell(row=i, column=2, value="Cliente").font = negrita12
            ws.cell(row=i, column=3, value="Nombre").font = negrita12
            ws.cell(row=i, column=4, value="Monto").font = negrita12
            ws.cell(row=i, column=5, value="Guía").font = negrita12
            ws.cell(row=i, column=6, value="Llegada").font = negrita12
            ws.cell(row=i, column=7, value="Llegada").font = negrita12
            ws.cell(row=i, column=8, value="Salida").font = negrita12
            ws.cell(row=i, column=9, value="Salida").font = negrita12
            ws.cell(row=i, column=10, value="Estado").font = negrita12
            ws.cell(row=i, column=11, value="Comentario").font = negrita12
            fecha_ctrl = wfecha
            sw = '1'

        if fecha_ctrl != wfecha:
            i+= 1
            nalmacen = Logalma.objects.get(almacen=walmacen).descripcion
            ncodigo = Logtral.objects.get(codigo=wcodigo).descripcion
            ws.cell(row=i, column=1, value="Fecha:").alignment = right_alignment
            ws.cell(row=i, column=1, value="Fecha:").font = negrita12
            ws.cell(row=i, column=2, value=wfecha).font = negrita12
            i+= 1
            ws.cell(row=i, column=1, value="Almacén:").alignment = right_alignment
            ws.cell(row=i, column=1, value="Almacén:").font = tamano12
            ws.cell(row=i, column=2, value=nalmacen).font = tamano12
            i+= 1
            ws.cell(row=i, column=1, value="Transacción:").alignment = right_alignment
            ws.cell(row=i, column=1, value="Transacción:").font = tamano12
            ws.cell(row=i, column=2, value=ncodigo)
            i+= 2

            ws.cell(row=i, column=6, value="Fecha").font = negrita12
            ws.cell(row=i, column=7, value="Hora").font = negrita12
            ws.cell(row=i, column=8, value="Fecha").font = negrita12
            ws.cell(row=i, column=9, value="Hora").font = negrita12
            i+= 1
            ws.cell(row=i, column=1, value="Documento").font = negrita12
            ws.cell(row=i, column=2, value="Cliente").font = negrita12
            ws.cell(row=i, column=3, value="Nombre").font = negrita12
            ws.cell(row=i, column=4, value="Monto").font = negrita12
            ws.cell(row=i, column=5, value="Guía").font = negrita12
            ws.cell(row=i, column=6, value="Llegada").font = negrita12
            ws.cell(row=i, column=7, value="Llegada").font = negrita12
            ws.cell(row=i, column=8, value="Salida").font = negrita12
            ws.cell(row=i, column=9, value="Salida").font = negrita12
            ws.cell(row=i, column=10, value="Estado").font = negrita12
            ws.cell(row=i, column=11, value="Comentario").font = negrita12
            fecha_ctrl = wfecha
        
        i+= 1
         
        if wstatus == "I":
            nwstatus = "Ingresada"
        else:
            if wstatus == "C":
                nwstatus = "Cancelada"
            else:
                if wstatus == "E":
                    nwstatus = "Entregado"
                else:
                    nwstatus = "Indefinido"

        ws.cell(row=i, column=1, value=wdocumento).alignment = center_alignment
        ws.cell(row=i, column=2, value=wcliente),
        ncliente = Crmclte.objects.get(cliente=wcliente).razon_social
        ws.cell(row=i, column=3, value=ncliente)
        ws.cell(row=i, column=4, value=wmonto).alignment = right_alignment
        ws.cell(row=i, column=4, value=wmonto).number_format = '$##,###,##0.00'
        ws.cell(row=i, column=5, value=wguia).alignment = right_alignment
        ws.cell(row=i, column=6, value=wfechaLlegada).alignment = center_alignment
        ws.cell(row=i, column=7, value=whoraLlegada).alignment = center_alignment
        ws.cell(row=i, column=8, value=wfechaEntrega).alignment = center_alignment
        ws.cell(row=i, column=9, value=whoraEntrega).alignment = center_alignment
        ws.cell(row=i, column=10, value=nwstatus).alignment = center_alignment
        ws.cell(row=i, column=11, value=wcomentario)

        logdemo = Logdemo.objects.filter(compania=wcompania, agencia=wagencia, fecha=wfecha,
                    almacen=walmacen, codigo=wcodigo, documento=wdocumento).values() #order_by('secuencia')

        i+= 1
        ws.cell(row=i, column=5, value="Ingreso").alignment = center_alignment
        ws.cell(row=i, column=5, value="Ingreso").font = negrita12
        ws.cell(row=i, column=6, value="Ingreso").alignment = center_alignment
        ws.cell(row=i, column=6, value="Ingreso").font = negrita12
        ws.cell(row=i, column=7, value="Entrega").alignment = center_alignment
        ws.cell(row=i, column=7, value="Entrega").font = negrita12
        ws.cell(row=i, column=8, value="Entrega").alignment = center_alignment
        ws.cell(row=i, column=8, value="Entrega").font = negrita12
        i+= 1
        ws.cell(row=i, column=2, value="Secuencia").alignment = right_alignment
        ws.cell(row=i, column=2, value="Secuencia").font = negrita12
        ws.cell(row=i, column=3, value="Cant. Pallets").alignment = right_alignment
        ws.cell(row=i, column=3, value="Cant. Pallets").font = negrita12
        ws.cell(row=i, column=4, value="No. Pallet").alignment = right_alignment
        ws.cell(row=i, column=4, value="No. Pallet").font = negrita12
        ws.cell(row=i, column=5, value="Cajas").alignment = right_alignment
        ws.cell(row=i, column=5, value="Cajas").font = negrita12
        ws.cell(row=i, column=6, value="Peso").alignment = right_alignment
        ws.cell(row=i, column=6, value="Peso").font = negrita12
        ws.cell(row=i, column=7, value="Cajas").alignment = center_alignment
        ws.cell(row=i, column=7, value="Cajas").font = negrita12
        ws.cell(row=i, column=8, value="Peso").alignment = center_alignment
        ws.cell(row=i, column=8, value="Peso").font = negrita12
        ws.cell(row=i, column=9, value="Monto").alignment = center_alignment
        ws.cell(row=i, column=9, value="Monto").font = negrita12
        ws.cell(row=i, column=10, value="Estado").alignment = center_alignment
        ws.cell(row=i, column=10, value="Estado").font = negrita12
        totDetalle = 0
        totCajas = 0
        totPeso = 0
        for demo in logdemo:
            wsecuencia = demo['secuencia']
            wcanpallets = demo['pallets']
            wnopallet = demo['orden_produccion']
            wcajas = demo['cajas']
            wpeso = demo['peso']
            wcajasdesp = demo['cajas_desp']
            wpesodesp = demo['peso_desp']
            wmonto = demo['monto']
            wstatusd = demo['status']

            if wmonto is None:
                wmonto = 0
            monto = float(wmonto)
            cajas = int(wcajas)
            peso = float(wpeso)
            totDetalle = totDetalle + monto
            totCajas = totCajas + cajas
            totPeso = totPeso + peso

            i+= 1
            if wcajas is None:
                wcajas = 0
            if wpeso is None:
                wpeso = 0
            if wcajasdesp is None:
                wcajasdesp = 0
            if wpesodesp is None:
                wpesodesp = 0

            ws.cell(row=i, column=2, value=wsecuencia).alignment = right_alignment
            ws.cell(row=i, column=3, value=wcanpallets).alignment = right_alignment
            ws.cell(row=i, column=4, value=wnopallet).alignment = center_alignment
            ws.cell(row=i, column=5, value=wcajas).alignment = right_alignment
            ws.cell(row=i, column=5, value=wcajas).number_format = '##.##0'
            ws.cell(row=i, column=6, value=wpeso).alignment = right_alignment
            ws.cell(row=i, column=6, value=wpeso).number_format = '##,##0.00'
            ws.cell(row=i, column=7, value=wcajasdesp).alignment = right_alignment
            ws.cell(row=i, column=7, value=wcajasdesp).number_format= '##,##0'
            ws.cell(row=i, column=8, value=wpesodesp).alignment = right_alignment
            ws.cell(row=i, column=8, value=wpesodesp).number_format = '##,##0.00'
            ws.cell(row=i, column=9, value=wmonto).alignment = right_alignment
            ws.cell(row=i, column=9, value=wmonto).number_format = '$##,##0.00'
            if wstatus == "I":
                nwstatus = "Ingresada"
            else:
                nwstatus = "Entregada"

            ws.cell(row=i, column=10, value=nwstatus).alignment = center_alignment

        i+= 1
        ws.cell(row=i, column=9, value=totDetalle).alignment = right_alignment
        ws.cell(row=i, column=9, value=totDetalle).font = negrita12
        ws.cell(row=i, column=9, value=totDetalle).number_format = '$##,##0.00'
        ws.cell(row=i, column=5, value=totCajas).alignment = right_alignment
        ws.cell(row=i, column=5, value=totCajas).font = negrita12
        ws.cell(row=i, column=5, value=totCajas).number_format = '##,##0'
        ws.cell(row=i, column=6, value=totPeso).alignment = right_alignment
        ws.cell(row=i, column=6, value=totPeso).font = negrita12
        ws.cell(row=i, column=6, value=totPeso).number_format = '###,##0.00'
        i+= 1

    #resultado = Caratvued.objects.filter(fecha=rfecha, compania=rcompania, matricula=rmatricula).aggregate(total=Sum('monto'))
    #monto = decimal.Decimal(resultado['total'])
                        
    # Adjust column widths
    column_widths = [12.64, 11, 16.55, 10, 7.27, 13, 13.82, 8.09, 9.45, 9.18, 21.45]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    id_passd = Segpaag.objects.get(compania=id_compania, agencia=id_agencia, aplicacion='CAR',
                            parametro='excel_passd').valor
    
    ws.protection.password = id_passd.strip()
    ws.protection.sheet = True
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="movimientos_ctofrio.xlsx"'
    wb.save(response)
    return response

@api_view(['POST'])
def excel_auxlogctmo(request):

    # Parsear el cuerpo de la solicitud
    try:
        vue_data = json.loads(request.body) if request.body else {}
        logger.info("Datos recibidos: %s", vue_data)  # Para depuración
    except json.JSONDecodeError:
        logger.error("Error al parsear JSON: %s", request.body)
        return Response({"error": "Invalid JSON data"}, status=400)

    # Validar que todos los campos de nivel raíz estén presentes
    required_fields = [
        "fecha", "codigo", "almacen", "documento", "status" ]
    missing_fields = [field for field in required_fields if field not in vue_data or vue_data[field] is None]
    if missing_fields:
        logger.error("Campos obligatorios faltantes: %s", missing_fields)
        return Response({"error": f"Missing required fields: {', '.join(missing_fields)}"}, status=400)

    # Extraer los campos
    ctmo_fecha = vue_data["fecha"]
    ctmo_codigo = vue_data["codigo"]
    ctmo_almacen = vue_data["almacen"]
    ctmo_documento = vue_data["documento"]
    ctmo_status = vue_data["status"]
    
    id_compania = '300'
    id_agencia = '001'
     
    now = datetime.now()
    w_fecha = now.date()
    w_hora = now.time()
    id_hora = w_hora.strftime('%H:%M:%S')

    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Auxiliar de Movimientos"

    # Set up fonts and alignment
    center_alignment = Alignment(horizontal="center")
    right_alignment = Alignment(horizontal="right")

    # Sección del encabezado
    ws.merge_cells('A1:J1')
    ws['A1'] = "COPADASA"
    ws['A1'].font = Font(name="Arial", size=16, bold=False)
    ws['A1'].alignment = center_alignment

    ws.merge_cells('A2:J2')
    ws['A2'] = "AUXILIAR DE MOVIMIENTOS"
    ws['A2'].font = Font(name="Arial", size=18, bold=True)
    ws['A2'].alignment = center_alignment

    ws['A3'] = "Fecha"
    ws['A3'].font = Font(name="Arial", size=13, bold=False)
    ws['A3'].alignment = right_alignment
    ws['B3'] = w_fecha
    ws['B3'].font = Font(name="Arial", size=13, bold=False)

    ws['A4'] = "Hora:"
    ws['A4'].font = Font(name="Arial", size=13, bold=False)
    ws['A4'].alignment = right_alignment
    ws['B4'] = w_hora
    ws['B4'].font = Font(name="Arial", size=13, bold=False)

    ws['A6'] = "Fecha:"
    ws['A6'].font = Font(name="Arial", size=13, bold=True)
    ws['A6'].alignment = right_alignment

    ws['B6'] = ctmo_fecha
    ws['B6'].font = Font(name="Arial", size=13, bold=True)

    logctmo = Logctmo.objects.filter(compania=id_compania, agencia=id_agencia, fecha=ctmo_fecha,
                    almacen=ctmo_almacen, codigo=ctmo_codigo, documento=ctmo_documento).values()
    
    for ctmo in logctmo:
        id_cliente = ctmo['cliente']
        comentario = ctmo['comentario']
        id_almacen = ctmo['almacen']
        estado = ctmo['status']
        guia = ctmo['guia_despacho']
        if estado == "D":
            fechaEntrega = ctmo['fecha_entrega']
            horaEntrega = ctmo['hora_entrega']
        
    if estado == "I":
        nestado = "Ingresado"
    else:
        if estado == "C":
            nestado = "Calculado"
        else:
            if estado == "D":
                nestado = "Despachado"
            else:
                nestado = "Indeterminado"

    ws['H6'] = "Estado:"
    ws['H6'].font = Font(name="Arial", size=13, bold=True)
    ws['H6'].alignment = right_alignment
    ws['I6'] = nestado
    ws['I6'].font = Font(name="Arial", size=13, bold=True)

    ws['A7'] = "Almacén:"
    ws['A7'].font = Font(name="Arial", size=13, bold=False)
    ws['A7'].alignment = right_alignment
    nalmacen = Logalma.objects.get(almacen=id_almacen).descripcion
    ws['B7'] = nalmacen
    ws['B7'].font = Font(name="Arial", size=13, bold=False)

    ws['H7'] = "Guía:"
    ws['H7'].font = Font(name="Arial", size=13, bold=False)
    ws['H7'].alignment = right_alignment 
    ws['I7'] = guia
    ws['I7'].font = Font(name="Arial", size=13, bold=False)
    ws['I7'].alignment = right_alignment

    ws['A8'] = "Transacción:"
    ws['A8'].font = Font(name="Arial", size=13, bold=False)
    ncodigo = Logtral.objects.get(codigo=ctmo_codigo).descripcion
    ws['B8'] = ncodigo
    ws['B8'].font = Font(name="Arial", size=13, bold=False)

   
    ws['A9'] = "Cliente:"
    ws['A9'].font = Font(name="Arial", size=13, bold=True)
    ws['A9'].alignment = right_alignment
    ws['B9'] = id_cliente
    ws['B9'].font = Font(name="Arial", size=13, bold=True)
    ncliente = Crmclte.objects.get(cliente=id_cliente).razon_social
    ws['C9'] = ncliente
    ws['C9'].font = Font(name="Arial", size=13, bold=True)

    if estado == "D":
        code_tr = Logtral.objects.get(accion='R', maneja_cliente='S').codigo
        logctmo_salida = Logctmo.objects.filter(compania=id_compania, agencia=id_agencia, 
                            almacen=id_almacen, codigo=code_tr, guia_despacho=guia, cliente=id_cliente).values()
        for salida in logctmo_salida:
            docto_salida = salida['documento']
            fechaEntrega = salida['fecha_entrega']
            horaEntrega = salida['hora_entrega']

    if estado == "D":
        ws['H9'] = "DESPACHO"
        ws['H9'].font = Font(name="Arial", size=13, bold=True)

    ws['A10'] = "Documento:"
    ws['A10'].font = Font(name="Arial", size=13, bold=False)
    ws['A10'].alignment = right_alignment
    ws['B10'] = docto_salida
    ws['B10'].font = Font(name="Arial", size=13, bold=False)
    ws['B10'].alignment = right_alignment

    if estado == "D":
        code_tr = Logtral.objects.get(accion='S', maneja_cliente='S').codigo
        logctmo_desp = Logctmo.objects.filter(compania=id_compania, agencia=id_agencia, 
                            almacen=id_almacen, codigo=code_tr, guia_despacho=guia, cliente=id_cliente).values()
        
        for desp in logctmo_desp:
            fechaLlegada = desp['fecha_llegada']
            horaLlegada = desp['hora_llegada']

    ws['A11'] = "Llegada:"
    ws['A11'].font = Font(name="Arial", size=12, bold=False)
    ws['A11'].alignment = right_alignment
    ws['B11'] = fechaLlegada
    ws['B11'].font = Font(name="Arial", size=13, bold=False)
    ws['C11'] = horaLlegada
    ws['C11'].font = Font(name="Arial", size=13, bold=False)

    if estado == "D":
        ws['H10'] = "Documento:"
        ws['H10'].font = Font(name="Arial", size=13, bold=False)
        ws['H10'].alignment = right_alignment
     
        ws['I10'] = docto_salida
        ws['I10'].font = Font(name="Arial", size=13, bold=False)
       

        ws['H11'] = "Fecha:"
        ws['H11'].font = Font(name="Arial", size=13, bold=False)
        ws['I11'] = fechaEntrega
        ws['I11'].font = Font(name="Arial", size=13, bold=False)
        ws['J11'] = horaEntrega
        ws['J11'].font = Font(name="Arial", size=13, bold=False)

    ws['A12'] = "Comentario:"
    ws['A12'].font = Font(name="Arial", size=13, bold=False)
    ws['A12'].alignment = right_alignment
    ws['B12'] = comentario
    ws['B12'].font = Font(name="Arial", size=13, bold=False)

    ws.merge_cells('A14:J14')
    ws['A14'] = "DETALLES"
    ws['A14'].font = Font(name="Arial", size=16, bold=True)
    ws['A14'].alignment = center_alignment

    ws['C15'] = "Entrada"
    ws['C15'].font = Font(name="Arial", size=12, bold=False)
    ws['D15'] = "Entrada"
    ws['D15'].font = Font(name="Arial", size=12, bold=False)
    ws['E15'] = "Entrada"
    ws['E15'].font = Font(name="Arial", size=12, bold=False)
    ws['F15'] = "Despacho"
    ws['F15'].font = Font(name="Arial", size=12, bold=False)
    ws['G15'] = "Despacho"
    ws['G15'].font = Font(name="Arial", size=12, bold=False)
    ws['H15'] = "Despacho"
    ws['H15'].font = Font(name="Arial", size=12, bold=False)

    ws['A16'] = "Secuencia"
    ws['A16'].font = Font(name="Arial", size=12, bold=False)
    ws['A16'].alignment = center_alignment
    ws['B16'] = "No. Pallet"
    ws['B16'].font = Font(name="Arial", size=12, bold=False )
    ws['B16'].alignment = center_alignment
    ws['C16'] = "Pallets"
    ws['C16'].font = Font(name="Arial", size=12, bold=False)
    ws['C16'].alignment = center_alignment
    ws['D16'] = "Cajas"
    ws['D16'].font = Font(name="Arial", size=12, bold=False)
    ws['D16'].alignment = right_alignment
    ws['E16'] = "Peso"
    ws['E16'].font = Font(name="Arial", size=12, bold=False)
    ws['E16'].alignment = right_alignment
    ws['F16'] = "Pallets"
    ws['F16'].font = Font(name="Arial", size=12, bold=False)
    ws['F16'].alignment = right_alignment
    ws['G16'] = "Cajas"
    ws['G16'].font = Font(name="Arial", size=12, bold=False)
    ws['G16'].alignment = right_alignment
    ws['H16'] = "Peso"
    ws['H16'].font = Font(name="Arial", size=12, bold=False)
    ws['I16'] = "Monto"
    ws['I16'].font = Font(name="Arial", size=12, bold=False)
    ws['I16'].alignment = right_alignment
    ws['J16'] = "Estado"
    ws['J16'].font = Font(name="Arial", size=12, bold=False)
    ws['J16'].alignment = center_alignment

    i = 17
    tamano12 = Font(size=12)
    tamano13 = Font(size=13)
    negrita12 = Font(size=12, bold=True)
    negrita13 = Font(size=13, bold=True)
    if estado == "D":
        code_tr = Logtral.objects.get(accion='S', maneja_cliente='S').codigo
        logctmo_desp = Logctmo.objects.filter(compania=id_compania, agencia=id_agencia, 
                            almacen=id_almacen, codigo=code_tr, guia_despacho=guia, cliente=id_cliente).values()
        
        for desp in logctmo_desp:
            id_cliente = desp['cliente']
            id_guia = desp['guia_despacho']
            id_fecha = desp['fecha']
            id_almacen = desp['almacen'] 
            id_documento = desp['documento']
            id_codigo = desp['codigo']

            
            logdemo = Logdemo.objects.filter(compania=id_compania, agencia=id_agencia, fecha=id_fecha,
                        almacen=id_almacen, codigo=id_codigo, documento=id_documento).values().order_by('secuencia')

            totMonto = 0
            totCajas = 0
            totPeso = 0
            totCajasdesp = 0
            totPesodesp = 0
            for demo in logdemo:
                wsecuencia = demo['secuencia']
                wcanpallets = demo['pallets']
                wnopallet = demo['orden_produccion']
                wcajas = 0
                wpeso = 0
                wmonto = 0
                 
                wcajas = demo['cajas']
                wpeso = demo['peso'] 
                wstatus = demo['status']

                if wmonto is None:
                    wmonto = 0

                monto = float(wmonto)
                cajas = int(wcajas)
                peso = float(wpeso)
                totMonto = totMonto + monto
                totCajas = totCajas + cajas
                totPeso = totPeso + peso

                ws.cell(row=i, column=1, value=wsecuencia).alignment = right_alignment
                ws.cell(row=i, column=2, value=wnopallet).alignment = right_alignment
                ws.cell(row=i, column=3, value=wcanpallets).alignment = center_alignment
                ws.cell(row=i, column=4, value=wcajas).alignment = right_alignment
                ws.cell(row=i, column=4, value=wcajas).number_format = '##,##0'
                ws.cell(row=i, column=5, value=wpeso).alignment = right_alignment
                ws.cell(row=i, column=5, value=wpeso).number_format = '##,##0.00'
                ws.cell(row=i, column=6, value=0).alignment = right_alignment
                ws.cell(row=i, column=6, value=0).number_format = '##0'
                ws.cell(row=i, column=7, value=0).alignment = right_alignment
                ws.cell(row=i, column=7, value=0).number_format = '##0'
                ws.cell(row=i, column=8, value=0).alignment = right_alignment
                ws.cell(row=i, column=8, value=0).number_format = '##0'
                ws.cell(row=i, column=9, value=0).alignment = right_alignment
                ws.cell(row=i, column=9, value=0).number_format = '$##,##0.00'
                if wstatus == "I":
                    nwstatus = "Ingresada"
                if wstatus == "C":
                    nwstatus = "Calculada"
                if wstatus == "D":
                    nwstatus = "Entregada"

                ws.cell(row=i, column=10, value=nwstatus).alignment = center_alignment
                i+= 1       
# Buscar Despacho si se realizó
        code_tr = Logtral.objects.get(accion='R', maneja_cliente='S').codigo
        logctmo_ent = Logctmo.objects.filter(compania=id_compania, agencia=id_agencia, almacen=id_almacen, 
                            codigo=code_tr, cliente=id_cliente, guia_despacho=id_guia).values()
        
        for ent in logctmo_ent:
            id_fecha = ent['fecha']
            id_almacen = ent['almacen']
            id_documento = ent['documento']

        logdemo_desp = Logdemo.objects.filter(compania=id_compania, agencia=id_agencia, 
                        fecha=id_fecha, almacen=id_almacen, codigo=code_tr, documento=id_documento).values().order_by('secuencia')
       
        for demo in logdemo_desp:
            wsecuencia = demo['secuencia']
            wcanpallets = demo['pallet_desp']
            wnopallet = demo['orden_produccion']
            wcajas = demo['cajas_desp']
            wpeso = demo['peso_desp'] 
            wmonto = demo['monto']
            wstatus = demo['status']
            
            if wmonto is None:
                wmonto = 0
            if wcajas is None:
                wcajas = 0
            if wcanpallets is None:
                wcanpallets = 0

            if wpeso is None:
                wpeso = 0

            wmonto = float(wmonto)
            wcajas = int(wcajas)
            wpeso = float(wpeso)
            totMonto = totMonto + wmonto
            totCajasdesp = totCajasdesp + wcajas
            totPesodesp = totPesodesp + wpeso

            ws.cell(row=i, column=1, value=wsecuencia).alignment = right_alignment
            ws.cell(row=i, column=2, value=wnopallet).alignment = right_alignment
            ws.cell(row=i, column=3, value=0).alignment = center_alignment
            ws.cell(row=i, column=4, value=0).alignment = right_alignment
            ws.cell(row=i, column=4, value=0).number_format = '##,##0'
            ws.cell(row=i, column=5, value=0).alignment = right_alignment
            ws.cell(row=i, column=5, value=0).number_format = '##,##0.00'
            ws.cell(row=i, column=6, value=wcanpallets).alignment = right_alignment
            ws.cell(row=i, column=6, value=wcanpallets).number_format = '##,##0'
            ws.cell(row=i, column=7, value=wcajas).alignment = right_alignment
            ws.cell(row=i, column=7, value=wcajas).number_format = '##,##0'
          
            ws.cell(row=i, column=8, value=wpeso).alignment = right_alignment
            ws.cell(row=i, column=8, value=wpeso).number_format = '##,##0.00'
            ws.cell(row=i, column=9, value=wmonto).alignment = right_alignment
            ws.cell(row=i, column=9, value=wmonto).number_format = '$###,##0.00'
            if wstatus == "I":
                nwstatus = "Ingresada"
            if wstatus == "C":
                nwstatus = "Calculada"
            if wstatus == "D":
                nwstatus = "Entregada"

            ws.cell(row=i, column=10, value=nwstatus).alignment = center_alignment
            i+= 1       
    else:
        logdemo = Logdemo.objects.filter(compania=id_compania, agencia=id_agencia, fecha=id_fecha,
                        almacen=ctmo_almacen, codigo=ctmo_codigo, documento=ctmo_documento).values().order_by('secuencia')

        totCajasdesp = 0
        totPesodesp = 0
        for demo in logdemo:
            wsecuencia = demo['secuencia']
            wcanpallets = demo['pallets']
            wnopallet = demo['orden_produccion']
            wcajas = 0
            wpeso = 0
                 
            wcajas = demo['cajas']
            wpeso = demo['peso'] 
            wmonto = demo['monto']
            wstatus = demo['status']

            if wmonto is None:
                wmonto = 0

            monto = float(wmonto)
            cajas = int(wcajas)
            peso = float(wpeso)
            totCajas = totCajas + cajas
            totPeso = totPeso + peso

            ws.cell(row=i, column=1, value=wsecuencia).alignment = right_alignment
            ws.cell(row=i, column=2, value=wnopallet).alignment = right_alignment
            ws.cell(row=i, column=3, value=wcanpallets).alignment = center_alignment
            ws.cell(row=i, column=4, value=wcajas).alignment = right_alignment
            ws.cell(row=i, column=4, value=wcajas).number_format = '##,##0'
            ws.cell(row=i, column=5, value=wpeso).alignment = right_alignment
            ws.cell(row=i, column=5, value=wpeso).number_format = '##,##0.00'
            ws.cell(row=i, column=6, value=0).alignment = right_alignment
            ws.cell(row=i, column=6, value=0).number_format = '##0'
            ws.cell(row=i, column=7, value=0).alignment = right_alignment
            ws.cell(row=i, column=7, value=0).number_format = '##0'
            ws.cell(row=i, column=8, value=0).alignment = right_alignment
            ws.cell(row=i, column=8, value=0).number_format = '##0'
            ws.cell(row=i, column=9, value=wmonto).alignment = right_alignment
            ws.cell(row=i, column=9, value=wmonto).number_format = '$##,##0.00'
            if wstatus == "I":
                nwstatus = "Ingresada"
            if wstatus == "C":
                nwstatus = "Calculada"
            if wstatus == "D":
                nwstatus = "Entregada"

            ws.cell(row=i, column=10, value=nwstatus).alignment = center_alignment
            i+= 1   

    ws.cell(row=i, column=1, value="Totales").font = negrita12
    ws.cell(row=i, column=4, value=totCajas).alignment = right_alignment
    ws.cell(row=i, column=4, value=totCajas).font = negrita12
    ws.cell(row=i, column=4, value=totCajas).number_format = '##,##0'
    ws.cell(row=i, column=5, value=totPeso).alignment = right_alignment
    ws.cell(row=i, column=5, value=totPeso).font = negrita12
    ws.cell(row=i, column=5, value=totPeso).number_format = '###,##0.00'
    ws.cell(row=i, column=6, value=0).alignment = right_alignment
    ws.cell(row=i, column=6, value=0).number_format = '##0'
    ws.cell(row=i, column=7, value=totCajasdesp).alignment = right_alignment
    ws.cell(row=i, column=7, value=totCajasdesp).number_format = '##,##0'
    ws.cell(row=i, column=7, value=totCajasdesp).font = negrita12
    ws.cell(row=i, column=8, value=totPesodesp).alignment = right_alignment
    ws.cell(row=i, column=8, value=totPesodesp).number_format = '##,##0.00'
    ws.cell(row=i, column=8, value=totPesodesp).font = negrita12
    ws.cell(row=i, column=9, value=totMonto).alignment = right_alignment
    ws.cell(row=i, column=9, value=totMonto).font = negrita12
    ws.cell(row=i, column=9, value=totMonto).number_format = '$##,##0.00'

    i+= 2
    ws.cell(row=i, column=1, value="Usuario Ingreso:").alignment = right_alignment
    ws.cell(row=i, column=2, value="contesis").font = negrita12
    i+=1
    ws.cell(row=i, column=1, value="Verificado Por:").alignment = right_alignment
    i+=1
    ws.cell(row=i, column=1, value="Supervisor:").alignment = right_alignment
     
    i+= 2
    ws.cell(row=i, column=1, value="View:").font = tamano12
    ws.cell(row=i, column=2, value="excel_auxlogctmo").font = tamano12
     
    #resultado = Caratvued.objects.filter(fecha=rfecha, compania=rcompania, matricula=rmatricula).aggregate(total=Sum('monto'))
    #monto = decimal.Decimal(resultado['total'])
                        
    # Adjust column widths
    column_widths = [12.64, 13, 10, 6.18, 6.64, 10.55, 9.55, 10, 10.45]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    id_passd = Segpaag.objects.get(compania=id_compania, agencia=id_agencia, aplicacion='CAR',
                            parametro='excel_passd').valor
    
    ws.protection.password = id_passd.strip()
    ws.protection.sheet = True
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="movimientos_ctofrio.xlsx"'
    wb.save(response)
    return response

@api_view(['POST'])
def excel_caratvue_resumen(request):

    # Parsear el cuerpo de la solicitud
    try:
        vue_data = json.loads(request.body) if request.body else {}
        logger.info("Datos recibidos: %s", vue_data)  # Para depuración
    except json.JSONDecodeError:
        logger.error("Error al parsear JSON: %s", request.body)
        return Response({"error": "Invalid JSON data"}, status=400)

    # Validar que todos los campos de nivel raíz estén presentes
    required_fields = [
        "fecha_inicial", "fecha_final", "compania" ]
    missing_fields = [field for field in required_fields if field not in vue_data or vue_data[field] is None]
    if missing_fields:
        logger.error("Campos obligatorios faltantes: %s", missing_fields)
        return Response({"error": f"Missing required fields: {', '.join(missing_fields)}"}, status=400)

    # Extraer los campos
    fecha_inicial = vue_data["fecha_inicial"]
    fecha_final = vue_data["fecha_final"]
    ae_compania = vue_data["compania"]
    
    id_compania = '300'
    id_agencia = '001'
     
    now = datetime.now()
    w_fecha = now.date()
    w_hora = now.time()
    id_hora = w_hora.strftime('%H:%M:%S')

    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Servicios"

    # Set up fonts and alignment
    center_alignment = Alignment(horizontal="center")
    right_alignment = Alignment(horizontal="right")

    # Sección del encabezado
    ws['A1'] = "Fecha:"
    ws['A1'].font = Font(name="Arial", size=10, bold=False)
    ws['A1'].alignment = right_alignment
    ws['B1'] = w_fecha
    ws['B1'].font = Font(name="Arial", size=10, bold=False)
    ws['B1'].alignment = right_alignment

    ws['A2'] = "Hora:"
    ws['A2'].font = Font(name="Arial", size=10, bold=False)
    ws['A2'].alignment = right_alignment
    ws['B2'] = w_hora
    ws['B2'].font = Font(name="Arial", size=10, bold=False)
    ws['B2'].alignment = right_alignment

    ws.merge_cells('C1:G1')
    ws['C1'] = "COPADASA"
    ws['C1'].font = Font(name="Arial", size=16, bold=False)
    ws['C1'].alignment = center_alignment

    ws.merge_cells('C2:G2')
    ws['C2'] = "RESUMEN DE SERVICIOS AÉREOS"
    ws['C2'].font = Font(name="Arial", size=16, bold=True)
    ws['C2'].alignment = center_alignment

    ws['C3'] = "Desde:"
    ws['C3'].font = Font(name="Arial", size=14, bold=True)
    ws['C3'].alignment = right_alignment
    ws['D3'] = fecha_inicial
    ws['D3'].font = Font(name="Arial", size=14, bold=True)
    
    ws['E3'] = "Hasta:"
    ws['E3'].font = Font(name="Arial", size=14, bold=True)
    ws['E3'].alignment = right_alignment
    ws['F3'] = fecha_final
    ws['F3'].font = Font(name="Arial", size=13, bold=True)

    ncompania = Carcoaer.objects.get(compania=ae_compania).nombre
    ws['A5'] = "Compañía:"
    ws['A5'].font = Font(name="Arial", size=12, bold=True)
    ws['A5'].alignment = right_alignment
    ws['B5'] = ae_compania
    ws['B5'].font = Font(name="Arial", size=12, bold=True)
    ws['C5'] = ncompania
    ws['C5'].font = Font(name="Arial", size=12, bold=True)

    ws['E7'] = "Fecha"
    ws['E7'].font = Font(name="Arial", size=12, bold=True)
    ws['E7'].alignment = center_alignment
    ws['F7'] = "Hora"
    ws['F7'].font = Font(name="Arial", size=12, bold=True)
    ws['F7'].alignment = center_alignment

    ws['A8'] = "Fecha"
    ws['A8'].font = Font(name="Arial", size=12, bold=True)
    ws['A8'].alignment = center_alignment
    ws['B8'] = "Aeronave"
    ws['B8'].font = Font(name="Arial", size=12, bold=True)
    ws['C8'] = "Descripción"
    ws['C8'].font = Font(name="Arial", size=12, bold=True)
    ws['D8'] = "Matrícula"
    ws['D8'].font = Font(name="Arial", size=12, bold=True)
    ws['E8'] = "Llegada"
    ws['E8'].font = Font(name="Arial", size=12, bold=True)
    ws['E8'].alignment = right_alignment
    ws['F8'] = "Llegada"
    ws['F8'].font = Font(name="Arial", size=12, bold=True)
    ws['F8'].alignment = right_alignment
    ws['G8'] = "Monto"
    ws['G8'].font = Font(name="Arial", size=12, bold=True)
    ws['G8'].alignment = right_alignment
    
# Encabezados
    caratvue = Caratvue.objects.filter(fecha__gte=fecha_inicial, fecha__lte=fecha_final,
                                       compania=ae_compania).values().order_by('fecha')
    
    tamano12 = Font(size=12)
    tamano13 = Font(size=13)
    negrita12 = Font(size=12, bold=True)
    negrita13 = Font(size=13, bold=True)

    i = 9
    totGeneral = 0
    for atvue in caratvue:
        id_fecha = atvue['fecha']
        id_aeronave = atvue['aeronave']
        id_matricula = atvue['matricula']
        fecha_llegada = atvue['fecha_llegada']
        hora_llegada = atvue['hora_llegada']
        monto = atvue['monto_servicio']

        naeronave = Cartiaero.objects.get(aeronave=id_aeronave).descripcion
        ws.cell(row=i, column=1, value=id_fecha).alignment = center_alignment
        ws.cell(row=i, column=2, value=id_aeronave).alignment = center_alignment
        ws.cell(row=i, column=3, value=naeronave)
        ws.cell(row=i, column=4, value=id_matricula)
        ws.cell(row=i, column=5, value=fecha_llegada)
        ws.cell(row=i, column=6, value=hora_llegada)
        ws.cell(row=i, column=7, value=monto).number_format = "$###,##0.00"
        i+= 1
        ws.cell(row=i, column=3, value="Cargo").alignment = center_alignment
        ws.cell(row=i, column=3, value="Cargo").font = negrita12
        ws.cell(row=i, column=4, value="Descripción").font = negrita12
        ws.cell(row=i, column=5, value="Tiempo").alignment = center_alignment
        ws.cell(row=i, column=5, value="Tiempo").font = negrita12
        ws.cell(row=i, column=6, value="Tarifa").alignment = right_alignment
        ws.cell(row=i, column=6, value="Tarifa").font = negrita12
        ws.cell(row=i, column=7, value="Monto").alignment = right_alignment
        ws.cell(row=i, column=7, value="Monto").font = negrita12

        totGeneral = totGeneral + monto 
        caratvued = Caratvued.objects.filter(fecha=id_fecha, compania=ae_compania, 
                            matricula=id_matricula).values().order_by('cargo')
        
        i+= 1
        totMonto = 0
        for detalle in caratvued:
            id_cargo = detalle['cargo']
            tiempo = detalle['tiempo_total']
            monto = detalle['monto']

            totMonto = totMonto + monto
            ncargo = Caratenvue.objects.get(cargo=id_cargo).nombre
            ws.cell(row=i, column=3, value=id_cargo).alignment = center_alignment
            ws.cell(row=i, column=3, value=id_cargo).font = tamano12
            ws.cell(row=i, column=4, value=ncargo).font = tamano12
            ws.cell(row=i, column=5, value=tiempo).alignment = center_alignment
            ws.cell(row=i, column=5, value=tiempo).font =tamano12
            ws.cell(row=i, column=7, value=monto).alignment = right_alignment
            ws.cell(row=i, column=7, value=monto).number_format = '$###,##0.00'
            i+= 1

        ws.cell(row=i, column=6, value="Total:").alignment = right_alignment
        ws.cell(row=i, column=6, value="Total:").font = negrita12
        ws.cell(row=i, column=7, value=totMonto).alignment = right_alignment
        ws.cell(row=i, column=7, value=totMonto).font = negrita12
        ws.cell(row=i, column=7, value=totMonto).number_format = '$###,##0.00'
        i+= 2
    
    ws.cell(row=i, column=6, value="Total Vigencia:").alignment = right_alignment
    ws.cell(row=i, column=6, value="Total Vigencia:").font = negrita13
    ws.cell(row=i, column=7, value=totGeneral).alignment = right_alignment
    ws.cell(row=i, column=7, value=totGeneral).number_format = '$##,###,##0.00'
    ws.cell(row=i, column=7, value=totGeneral).font = negrita13
    i+= 2
    ws.cell(row=i, column=1, value="View:").font = tamano12
    ws.cell(row=i, column=2, value="excel_caratvue_resumen").font = tamano12

    # Adjust column widths
    column_widths = [11.36, 10.18, 17.55, 15.91, 10.73, 10.27, 10.91]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    id_passd = Segpaag.objects.get(compania=id_compania, agencia=id_agencia, aplicacion='CAR',
                            parametro='excel_passd').valor
    
    ws.protection.password = id_passd.strip()
    ws.protection.sheet = True 
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="movimientos_ctofrio.xlsx"'
    wb.save(response)
    return response
